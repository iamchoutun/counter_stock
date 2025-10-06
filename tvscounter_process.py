from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from collections import defaultdict
from typing import Dict

def tvscounter_process(input_file: str, output_file=None) -> Dict[str, Dict[str, Dict[str, int]]]:
    """
    Process a TVS Excel file to count Good and Defect items per model.

    Args:
        input_file (str): Path to the TVS Excel file.
        output_file (str, optional): Placeholder for compatibility, not used.

    Returns:
        Dict[str, Dict[str, Dict[str, int]]]: Dictionary of counts per model and category.
            Example: {
                "Hybrid": {"C-HD ATV": {"Good": 2, "Defect": 1}, ...},
                "SKWAMX3 : TRUEIDTVGEN2 SKY TICC": {"Good": 10, "Defect": 0}, ...
            }
    """
    # Load workbook and get active sheet
    wb = load_workbook(input_file, data_only=True)
    ws = wb.active

    # Convert Excel column letters to 0-based indices
    tvs_model_col = column_index_from_string("C") - 1  # Model
    tvs_status_col = column_index_from_string("F") - 1  # Status

    # Initialize summary dictionary
    summary = {
        "Hybrid": defaultdict(lambda: {"Good": 0, "Defect": 0}),
        "SKWAMX3 : TRUEIDTVGEN2 SKY TICC": {"Good": 0, "Defect": 0},
        "SKWAMX5M : SMARTTRUEIDTVGEN3 T3 SKY TICC": {"Good": 0, "Defect": 0},
        "SKWAMX5M-NO : SMARTTRUETDTVGEN3.1 T3 SKY TICC": {"Good": 0, "Defect": 0}
    }

    hybrid_models = [
        "C-HD ATV : SMC HD SKY TVG CATV",
        "HBSK-NET100C : STD HYBRID SKY100 TVG"
    ]

    predefined_models = {
        "SKWAMX3 : TRUEIDTVGEN2 SKY TICC",
        "SKWAMX5M : SMARTTRUEIDTVGEN3 T3 SKY TICC",
        "SKWAMX5M-NO : SMARTTRUETDTVGEN3.1 T3 SKY TICC"
    }

    # Iterate through each row starting from row 2
    for row in ws.iter_rows(min_row=2, values_only=True):
        model = row[tvs_model_col]
        status = str(row[tvs_status_col] or "").strip().upper()
        is_defect = status == "DEFECT"

        if model in hybrid_models:
            summary["Hybrid"][model]["Defect" if is_defect else "Good"] += 1
        elif model in predefined_models:
            summary[model]["Defect" if is_defect else "Good"] += 1
        # Any other models are ignored

    # Calculate total for Hybrid
    total_good = sum(v["Good"] for v in summary["Hybrid"].values())
    total_defect = sum(v["Defect"] for v in summary["Hybrid"].values())
    summary["Hybrid"]["Total"] = {"Good": total_good, "Defect": total_defect}

    return summary
