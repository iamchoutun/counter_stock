from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from typing import Dict

def tolcounter_process(input_file: str, output_file=None) -> Dict[str, Dict[str, int]]:
    """
    Process a TOL Excel file to count Good and Defect items per model.

    Args:
        input_file (str): Path to the TOL Excel file.
        output_file (str, optional): Placeholder for compatibility, not used.

    Returns:
        Dict[str, Dict[str, int]]: Dictionary of counts per model.
            Example: {"T626Pro": {"Good": 10, "Defect": 2}}
    """
    # Load workbook and get active sheet
    wb = load_workbook(input_file, data_only=True)
    ws = wb.active

    # Convert Excel column letters to 0-based indices
    tol_model_col = column_index_from_string("I") - 1  # Model
    tol_status_col = column_index_from_string("M") - 1  # Status
    tol_dongle_col = column_index_from_string("H") - 1  # Dongle flag

    summary: Dict[str, Dict[str, int]] = {}

    # Iterate through each row starting from row 2
    for row in ws.iter_rows(min_row=2, values_only=True):
        model = row[tol_model_col]
        status = row[tol_status_col]
        dongle_flag = row[tol_dongle_col]

        # Separate T626ProV2 with Dongle flag
        if model == "T626ProV2":
            if dongle_flag and "dongle" in str(dongle_flag).lower():
                model_key = "T626ProV2_Dongle"
            else:
                model_key = "T626ProV2"
        else:
            model_key = model

        # Initialize dictionary if model not in summary
        if model_key not in summary:
            summary[model_key] = {"Good": 0, "Defect": 0}

        # Count Good or Defect
        if status and "DEFECT" in str(status).upper():
            summary[model_key]["Defect"] += 1
        else:
            summary[model_key]["Good"] += 1

    return summary
